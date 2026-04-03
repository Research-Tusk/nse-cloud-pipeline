"""
Excel Sync — Append new daily records from Supabase to OneDrive Excel workbook.
Runs as part of the GitHub Actions exchange-update workflow.

Flow:
  1. Authenticate with Azure AD (client_credentials) -> access token
  2. Download existing Excel from OneDrive (or create new if not found)
  3. Fetch all NSE + BSE daily records from Supabase
  4. Append only new dates (dedup by column A)
  5. Upload modified workbook back to OneDrive

Environment variables:
  SUPABASE_URL, SUPABASE_KEY         — Supabase credentials (existing)
  MSGRAPH_TENANT_ID                  — Azure AD tenant ID
  MSGRAPH_CLIENT_ID                  — App registration client ID
  MSGRAPH_CLIENT_SECRET              — App registration client secret
  ONEDRIVE_USER                      — UPN of the OneDrive owner (e.g. research@tuskinvest.com)
  EXCEL_FILE_PATH                    — Path within OneDrive, e.g. /NSE_BSE_Daily.xlsx
"""

import io
import os
import sys
import time
from datetime import datetime

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# -- Config --
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

MSGRAPH_TENANT_ID = os.environ.get("MSGRAPH_TENANT_ID", "")
MSGRAPH_CLIENT_ID = os.environ.get("MSGRAPH_CLIENT_ID", "")
MSGRAPH_CLIENT_SECRET = os.environ.get("MSGRAPH_CLIENT_SECRET", "")
ONEDRIVE_USER = os.environ.get("ONEDRIVE_USER", "")
EXCEL_FILE_PATH = os.environ.get("EXCEL_FILE_PATH", "/NSE_BSE_Daily.xlsx")

GRAPH_BASE = "https://graph.microsoft.com/v1.0"

# Column definitions (excluding created_at, updated_at metadata)
NSE_COLUMNS = [
    "date", "day", "fy_quarter", "fy_month", "fy",
    "if_turnover", "sf_turnover", "io_notional", "io_premium",
    "so_notional", "so_premium", "io_pcr",
    "total_contracts", "total_turnover",
    "if_rev", "sf_rev", "fut_rev", "io_rev", "so_rev", "opt_rev", "fo_rev",
    "cash_traded_value", "cash_rev", "total_rev", "pn_ratio", "vix",
]

BSE_COLUMNS = [
    "date", "day", "fy_quarter", "fy_month", "fy",
    "total_contracts", "total_turnover", "futures_turnover",
    "options_notional_turnover", "options_premium_turnover",
    "fut_rev", "opt_rev", "fo_rev",
    "cash_traded_value", "cash_securities", "cash_trades", "cash_quantity",
    "cash_rev", "total_rev", "pn_ratio",
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center")


# ================================================================
# SECTION 1: Microsoft Graph API
# ================================================================

def get_graph_token():
    """Acquire access token via OAuth2 client_credentials flow."""
    url = f"https://login.microsoftonline.com/{MSGRAPH_TENANT_ID}/oauth2/v2.0/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": MSGRAPH_CLIENT_ID,
        "client_secret": MSGRAPH_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
    }
    resp = requests.post(url, data=data, timeout=30)
    if resp.status_code != 200:
        print(f"ERROR: Auth failed ({resp.status_code}): {resp.text}")
        sys.exit(1)
    token = resp.json().get("access_token")
    if not token:
        print(f"ERROR: No access_token in response: {resp.json()}")
        sys.exit(1)
    print("Authenticated with Microsoft Graph.")
    return token


def _graph_headers(token):
    return {"Authorization": f"Bearer {token}"}


def _drive_item_url():
    """Build the Graph API URL for the target Excel file.
    Uses /users/{upn}/drive for OneDrive for Business (app-only auth)."""
    return f"{GRAPH_BASE}/users/{ONEDRIVE_USER}/drive/root:{EXCEL_FILE_PATH}:"


def download_excel(token):
    """Download the Excel file from OneDrive. Returns bytes or None if not found."""
    url = f"{_drive_item_url()}/content"
    resp = requests.get(url, headers=_graph_headers(token), timeout=60)
    if resp.status_code == 200:
        print(f"Downloaded existing Excel ({len(resp.content)} bytes).")
        return resp.content
    if resp.status_code == 404:
        print("Excel file not found on OneDrive -- will create a new one.")
        return None
    print(f"ERROR: Download failed ({resp.status_code}): {resp.text}")
    sys.exit(1)


def upload_excel(token, file_bytes, retry=True):
    """Upload (create or replace) the Excel file on OneDrive."""
    url = f"{_drive_item_url()}/content"
    headers = {**_graph_headers(token), "Content-Type": "application/octet-stream"}
    resp = requests.put(url, headers=headers, data=file_bytes, timeout=60)
    if resp.status_code in (200, 201):
        print(f"Uploaded Excel ({len(file_bytes)} bytes) to OneDrive.")
        return
    if resp.status_code == 429 and retry:
        wait = int(resp.headers.get("Retry-After", 5))
        print(f"Rate limited -- waiting {wait}s before retry...")
        time.sleep(wait)
        return upload_excel(token, file_bytes, retry=False)
    print(f"ERROR: Upload failed ({resp.status_code}): {resp.text}")
    if retry:
        print("Retrying upload in 2s...")
        time.sleep(2)
        return upload_excel(token, file_bytes, retry=False)
    sys.exit(1)


# ================================================================
# SECTION 2: Supabase
# ================================================================

def get_supabase_client():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def fetch_all_daily(supabase, table):
    """Fetch all records from a daily table, paginated."""
    all_records = []
    offset = 0
    batch_size = 1000
    while True:
        result = (
            supabase.table(table)
            .select("*")
            .order("date")
            .range(offset, offset + batch_size - 1)
            .execute()
        )
        if not result.data:
            break
        all_records.extend(result.data)
        if len(result.data) < batch_size:
            break
        offset += batch_size
    return all_records


# ================================================================
# SECTION 3: Excel workbook operations
# ================================================================

def _format_header_row(ws, columns):
    """Apply header formatting to row 1."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def create_new_workbook():
    """Create a fresh workbook with NSE Daily and BSE Daily sheets."""
    wb = Workbook()

    # NSE Daily sheet (rename default sheet)
    ws_nse = wb.active
    ws_nse.title = "NSE Daily"
    _format_header_row(ws_nse, NSE_COLUMNS)

    # BSE Daily sheet
    ws_bse = wb.create_sheet("BSE Daily")
    _format_header_row(ws_bse, BSE_COLUMNS)

    print("Created new workbook with NSE Daily and BSE Daily sheets.")
    return wb


def get_existing_dates(ws):
    """Read date values from column A, row 2 onward."""
    dates = set()
    for row in range(2, (ws.max_row or 1) + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        if isinstance(val, datetime):
            dates.add(val.strftime("%Y-%m-%d"))
        else:
            dates.add(str(val).strip())
    return dates


def append_rows(ws, records, columns):
    """Append records as new rows. Returns count of rows added."""
    existing = get_existing_dates(ws)
    new_records = [r for r in records if r.get("date") not in existing]
    new_records.sort(key=lambda x: x["date"])

    for rec in new_records:
        row = [rec.get(col) for col in columns]
        ws.append(row)

    return len(new_records)


def ensure_sheet(wb, name, columns):
    """Get or create a sheet by name with headers."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    _format_header_row(ws, columns)
    print(f"Created missing sheet: {name}")
    return ws


# ================================================================
# SECTION 4: Main
# ================================================================

def main():
    # Validate required env vars
    missing = []
    for var in ["SUPABASE_URL", "SUPABASE_KEY", "MSGRAPH_TENANT_ID",
                "MSGRAPH_CLIENT_ID", "MSGRAPH_CLIENT_SECRET", "ONEDRIVE_USER"]:
        if not os.environ.get(var):
            missing.append(var)
    if missing:
        print(f"ERROR: Missing environment variables: {', '.join(missing)}")
        sys.exit(1)

    # 1. Authenticate with Microsoft Graph
    token = get_graph_token()

    # 2. Download existing Excel or create new
    content = download_excel(token)
    if content:
        wb = load_workbook(io.BytesIO(content))
    else:
        wb = create_new_workbook()

    # 3. Fetch all daily records from Supabase
    print("Connecting to Supabase...")
    supabase = get_supabase_client()

    print("Fetching NSE daily records...")
    nse_records = fetch_all_daily(supabase, "nse_daily")
    print(f"  Fetched {len(nse_records)} NSE records from Supabase.")

    print("Fetching BSE daily records...")
    bse_records = fetch_all_daily(supabase, "bse_daily")
    print(f"  Fetched {len(bse_records)} BSE records from Supabase.")

    # 4. Append new rows to each sheet
    ws_nse = ensure_sheet(wb, "NSE Daily", NSE_COLUMNS)
    nse_added = append_rows(ws_nse, nse_records, NSE_COLUMNS)

    ws_bse = ensure_sheet(wb, "BSE Daily", BSE_COLUMNS)
    bse_added = append_rows(ws_bse, bse_records, BSE_COLUMNS)

    print(f"\nNew rows: NSE={nse_added}, BSE={bse_added}")

    # 5. Upload if anything changed
    if nse_added == 0 and bse_added == 0:
        print("No new rows to sync. Skipping upload.")
        return

    buf = io.BytesIO()
    wb.save(buf)
    upload_excel(token, buf.getvalue())

    print(f"Done. Synced {nse_added + bse_added} new rows to OneDrive.")


if __name__ == "__main__":
    main()
